"""Statement import orchestration service handling two-step preview & confirmation workflow."""

import os
import io
import csv
import json
import openpyxl
import uuid
import hashlib
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional, Tuple
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.statement_import import StatementImport
from app.models.bank_transaction import BankTransaction
from app.repositories.statement_import_repository import StatementImportRepository
from app.repositories.bank_transaction_repository import BankTransactionRepository
from app.services.statement_parser_service import StatementParserService
from app.schemas.statement_import import (
    BankTransactionListResponse,
    BankTransactionResponse,
    HeaderItem,
    ImportConfirmRequest,
    ImportPreviewResponse,
    ImportSummaryResponse,
    StatementColumnMapping,
    StatementImportDetailResponse,
    StatementImportListItemResponse,
    StatementImportListResponse,
)

# In-memory preview token metadata store (token -> dict)
_PREVIEW_SESSIONS: Dict[str, Dict[str, Any]] = {}


class StatementImportService:
    """Service orchestrating statement import previews, confirmations, deduplication, and persistence."""

    def __init__(
        self,
        parser_service: Optional[StatementParserService] = None,
        import_repo: Optional[StatementImportRepository] = None,
        txn_repo: Optional[BankTransactionRepository] = None,
    ):
        self.parser = parser_service or StatementParserService()
        self.import_repo = import_repo or StatementImportRepository()
        self.txn_repo = txn_repo or BankTransactionRepository()

    def preview_import(
        self,
        file_bytes: bytes,
        filename: str,
        admin_user_id: int,
        sheet_name: Optional[str] = None,
        header_row_index: int = 1,
    ) -> ImportPreviewResponse:
        """Step 1: Inspect uploaded statement, extract headers/sheets/preview rows. Zero DB writes."""

        file_type, available_sheets, selected_sheet_name, headers, preview_rows, total_rows = self.parser.inspect_file(
            file_bytes=file_bytes,
            filename=filename,
            sheet_name=sheet_name,
            header_row_index=header_row_index,
        )

        checksum = self.parser.calculate_checksum(file_bytes)
        preview_token = f"prev_sec_{uuid.uuid4().hex}"
        expires_at = datetime.now(timezone.utc) + timedelta(minutes=30)

        # Store preview session metadata
        _PREVIEW_SESSIONS[preview_token] = {
            "admin_user_id": admin_user_id,
            "expires_at": expires_at,
            "filename": filename,
            "file_type": file_type,
            "file_bytes": file_bytes,
            "file_checksum_sha256": checksum,
            "available_sheets": available_sheets,
            "selected_sheet_name": selected_sheet_name,
            "header_row_index": header_row_index,
        }

        return ImportPreviewResponse(
            preview_token=preview_token,
            filename=filename,
            file_type=file_type,
            file_size=len(file_bytes),
            file_checksum_sha256=checksum,
            available_sheets=available_sheets,
            selected_sheet_name=selected_sheet_name,
            header_row_index=header_row_index,
            headers=headers,
            preview_rows=preview_rows,
            total_detected_rows=total_rows,
            expires_in_seconds=1800,
        )

    async def confirm_import(
        self,
        db: AsyncSession,
        admin_user_id: int,
        payload: ImportConfirmRequest,
    ) -> ImportSummaryResponse:
        """Step 2: Confirm import, validate position mapping, check exact-file idempotency, deduplicate txns, persist."""

        # 1. Validate preview token session & ownership
        session = _PREVIEW_SESSIONS.get(payload.preview_token)
        if not session:
            raise ValueError("Import preview has expired or is invalid. Please upload the file again.")

        if session["admin_user_id"] != admin_user_id:
            raise ValueError("Unauthorized access to preview session.")

        if datetime.now(timezone.utc) > session["expires_at"]:
            _PREVIEW_SESSIONS.pop(payload.preview_token, None)
            raise ValueError("Import preview has expired. Please upload the file again.")

        file_bytes: bytes = session["file_bytes"]
        filename: str = session["filename"]
        file_type: str = session["file_type"]
        file_checksum: str = session["file_checksum_sha256"]
        sheet_name: Optional[str] = payload.sheet_name or session["selected_sheet_name"]
        header_row_index: int = payload.header_row_index
        source_tz: str = payload.source_timezone or settings.STATEMENT_SOURCE_TIMEZONE

        mapping = payload.column_mapping

        # 2. Compute canonical mapping hash for exact duplicate lookup
        mapping_dict = mapping.model_dump()
        canonical_raw = f"{file_checksum}|{sheet_name or ''}|{header_row_index}|{json.dumps(mapping_dict, sort_keys=True)}"
        canonical_hash = hashlib.sha256(canonical_raw.encode("utf-8")).hexdigest()

        # 3. Check exact-file idempotency
        existing_import = await self.import_repo.get_by_canonical_hash(db, canonical_hash)
        if existing_import:
            # Single-use invalidation
            _PREVIEW_SESSIONS.pop(payload.preview_token, None)
            return ImportSummaryResponse(
                already_imported=True,
                import_public_id=existing_import.public_id,
                filename=existing_import.filename,
                file_type=existing_import.file_type,
                selected_sheet_name=existing_import.selected_sheet_name,
                status=existing_import.status,
                total_rows=existing_import.total_rows,
                valid_rows=existing_import.valid_rows,
                invalid_rows=existing_import.invalid_rows,
                duplicate_rows=existing_import.duplicate_rows,
                new_transactions=existing_import.new_transactions,
                rows_without_reference=existing_import.rows_without_reference,
                created_at=existing_import.created_at,
                completed_at=existing_import.completed_at,
                error_summary=existing_import.error_summary,
                message="This statement file with identical sheet and column mapping has already been imported.",
            )

        # 4. Parse all rows from file
        headers, data_rows = self._read_all_rows(file_bytes, file_type, sheet_name, header_row_index)

        # Validate column indices exist in source file
        max_idx = max(
            mapping.reference_id.column_index,
            mapping.amount.column_index,
            mapping.transaction_at.column_index if mapping.transaction_at else 0,
            mapping.direction.column_index if mapping.direction else 0,
            mapping.utr.column_index if mapping.utr else 0,
            mapping.counterparty_name.column_index if mapping.counterparty_name else 0,
            mapping.description.column_index if mapping.description else 0,
        )

        total_rows = len(data_rows)
        valid_rows = 0
        invalid_rows = 0
        rows_without_reference = 0

        parsed_candidates: List[Dict[str, Any]] = []
        error_summary_list: List[Dict[str, Any]] = []

        for row_idx, row in enumerate(data_rows, start=header_row_index + 1):
            if len(row) <= max_idx and not any(row):
                # Empty trailing row, skip
                continue

            try:
                # Extract values by 0-based column_index
                ref_val = row[mapping.reference_id.column_index] if mapping.reference_id.column_index < len(row) else None
                amt_val = row[mapping.amount.column_index] if mapping.amount.column_index < len(row) else None

                date_val = None
                if mapping.transaction_at and mapping.transaction_at.column_index < len(row):
                    date_val = row[mapping.transaction_at.column_index]

                dir_val = None
                if mapping.direction and mapping.direction.column_index < len(row):
                    dir_val = row[mapping.direction.column_index]

                utr_val = None
                if mapping.utr and mapping.utr.column_index < len(row):
                    utr_val = row[mapping.utr.column_index]

                cp_val = None
                if mapping.counterparty_name and mapping.counterparty_name.column_index < len(row):
                    cp_val = row[mapping.counterparty_name.column_index]

                desc_val = None
                if mapping.description and mapping.description.column_index < len(row):
                    desc_val = row[mapping.description.column_index]

                # Normalize fields
                norm_ref = self.parser.normalize_reference_id(ref_val)
                norm_amt = self.parser.parse_amount(amt_val)
                norm_date = self.parser.parse_date(date_val, default_tz_str=source_tz)
                norm_dir = self.parser.parse_direction(dir_val)
                norm_utr = self.parser.normalize_reference_id(utr_val)
                norm_cp = str(cp_val).strip() if cp_val is not None and str(cp_val).strip() else None
                norm_desc = str(desc_val).strip() if desc_val is not None and str(desc_val).strip() else None

                if norm_amt is None:
                    raise ValueError(f"Missing or unparseable amount in row {row_idx}.")

                if norm_ref is None:
                    rows_without_reference += 1
                    continue

                source_key = self.parser.compute_source_transaction_key(
                    source="GOOGLE_PAY",
                    reference_id=norm_ref,
                    utr=norm_utr,
                    amount_inr=norm_amt,
                    direction=norm_dir,
                    transaction_at=norm_date,
                    counterparty=norm_cp,
                    description=norm_desc,
                )

                valid_rows += 1
                parsed_candidates.append({
                    "row_idx": row_idx,
                    "transaction_at": norm_date,
                    "amount_inr": norm_amt,
                    "direction": norm_dir,
                    "reference_id": norm_ref,
                    "utr": norm_utr,
                    "counterparty_name": norm_cp,
                    "description": norm_desc,
                    "source": "GOOGLE_PAY",
                    "source_transaction_key": source_key,
                    "raw_row_data": {str(k): str(v) for k, v in enumerate(row)},
                })

            except Exception as exc:
                invalid_rows += 1
                if len(error_summary_list) < 100:
                    error_summary_list.append({
                        "row_number": row_idx,
                        "field": "general",
                        "message": str(exc),
                    })

        # 5. Deduplicate candidates against existing DB keys
        candidate_keys = {c["source_transaction_key"] for c in parsed_candidates}
        existing_db_keys = await self.txn_repo.get_existing_keys(db, "GOOGLE_PAY", candidate_keys)

        new_txns_to_create: List[BankTransaction] = []
        duplicate_rows = 0
        seen_batch_keys: set = set()

        # Create StatementImport entity first
        import_entity = StatementImport(
            filename=filename,
            file_type=file_type,
            file_size=len(file_bytes),
            file_checksum_sha256=file_checksum,
            canonical_mapping_hash=canonical_hash,
            source="GOOGLE_PAY",
            selected_sheet_name=sheet_name,
            header_row_index=header_row_index,
            column_mapping=mapping_dict,
            status="COMPLETED" if invalid_rows == 0 else "COMPLETED_WITH_ERRORS",
            total_rows=total_rows,
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
            duplicate_rows=0,
            new_transactions=0,
            rows_without_reference=rows_without_reference,
            error_summary={"errors": error_summary_list, "total_errors": len(error_summary_list)} if error_summary_list else None,
            imported_by=admin_user_id,
            completed_at=datetime.now(timezone.utc),
        )

        db_import = await self.import_repo.create(db, import_entity)

        for c in parsed_candidates:
            s_key = c["source_transaction_key"]
            if s_key in existing_db_keys or s_key in seen_batch_keys:
                duplicate_rows += 1
            else:
                seen_batch_keys.add(s_key)
                new_txns_to_create.append(
                    BankTransaction(
                        statement_import_id=db_import.id,
                        transaction_at=c["transaction_at"],
                        amount_inr=c["amount_inr"],
                        direction=c["direction"],
                        reference_id=c["reference_id"],
                        utr=c["utr"],
                        counterparty_name=c["counterparty_name"],
                        description=c["description"],
                        source=c["source"],
                        source_transaction_key=s_key,
                        raw_row_data=c["raw_row_data"],
                    )
                )

        db_import.duplicate_rows = duplicate_rows
        db_import.new_transactions = len(new_txns_to_create)

        # Bulk insert new transactions
        await self.txn_repo.bulk_create(db, new_txns_to_create)
        await db.commit()

        # Single-use invalidation
        _PREVIEW_SESSIONS.pop(payload.preview_token, None)

        return ImportSummaryResponse(
            already_imported=False,
            import_public_id=db_import.public_id,
            filename=db_import.filename,
            file_type=db_import.file_type,
            selected_sheet_name=db_import.selected_sheet_name,
            status=db_import.status,
            total_rows=db_import.total_rows,
            valid_rows=db_import.valid_rows,
            invalid_rows=db_import.invalid_rows,
            duplicate_rows=db_import.duplicate_rows,
            new_transactions=db_import.new_transactions,
            rows_without_reference=db_import.rows_without_reference,
            created_at=db_import.created_at,
            completed_at=db_import.completed_at,
            error_summary=db_import.error_summary,
            message="Statement import completed successfully.",
        )

    def _read_all_rows(
        self, file_bytes: bytes, file_type: str, sheet_name: Optional[str], header_row_index: int
    ) -> Tuple[List[str], List[List[Any]]]:
        """Read all rows from CSV or XLSX file."""
        if file_type == "csv":
            content_str = file_bytes.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(content_str))
            all_rows = list(reader)
            headers = [c.strip() for c in all_rows[header_row_index - 1]]
            return headers, all_rows[header_row_index:]
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            target_sheet = sheet_name or wb.sheetnames[0]
            ws = wb[target_sheet]
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            headers = [str(c).strip() if c is not None else "" for c in all_rows[header_row_index - 1]]
            return headers, all_rows[header_row_index:]

    async def list_imports(
        self, db: AsyncSession, page: int = 1, page_size: int = 20
    ) -> StatementImportListResponse:
        """Fetch paginated import history."""
        items, total = await self.import_repo.list_paginated(db, page, page_size)
        total_pages = (total + page_size - 1) // page_size if total > 0 else 1

        list_items = [
            StatementImportListItemResponse(
                public_id=imp.public_id,
                filename=imp.filename,
                file_type=imp.file_type,
                source=imp.source,
                selected_sheet_name=imp.selected_sheet_name,
                status=imp.status,
                total_rows=imp.total_rows,
                valid_rows=imp.valid_rows,
                duplicate_rows=imp.duplicate_rows,
                new_transactions=imp.new_transactions,
                rows_without_reference=imp.rows_without_reference,
                imported_by_name=admin_name,
                created_at=imp.created_at,
            )
            for imp, admin_name in items
        ]

        return StatementImportListResponse(
            items=list_items,
            total=total,
            page=page,
            page_size=page_size,
            total_pages=total_pages,
        )

    async def get_import_detail(
        self, db: AsyncSession, import_public_id: uuid.UUID
    ) -> StatementImportDetailResponse:
        """Fetch single import detail record."""
        res = await self.import_repo.get_by_public_id(db, import_public_id)
        if not res:
            raise ValueError(f"Statement import '{import_public_id}' not found.")

        imp, admin_name = res

        return StatementImportDetailResponse(
            public_id=imp.public_id,
            filename=imp.filename,
            file_type=imp.file_type,
            file_size=imp.file_size,
            file_checksum_sha256=imp.file_checksum_sha256,
            canonical_mapping_hash=imp.canonical_mapping_hash,
            source=imp.source,
            selected_sheet_name=imp.selected_sheet_name,
            header_row_index=imp.header_row_index,
            column_mapping=imp.column_mapping,
            status=imp.status,
            total_rows=imp.total_rows,
            valid_rows=imp.valid_rows,
            invalid_rows=imp.invalid_rows,
            duplicate_rows=imp.duplicate_rows,
            new_transactions=imp.new_transactions,
            rows_without_reference=imp.rows_without_reference,
            error_summary=imp.error_summary,
            imported_by_name=admin_name,
            created_at=imp.created_at,
            completed_at=imp.completed_at,
        )

    async def list_import_transactions(
        self, db: AsyncSession, import_public_id: uuid.UUID, page: int = 1, page_size: int = 20
    ) -> BankTransactionListResponse:
        """Fetch paginated bank transactions for a completed import."""
        res = await self.import_repo.get_by_public_id(db, import_public_id)
        if not res:
            raise ValueError(f"Statement import '{import_public_id}' not found.")

        imp, _ = res
        items, total = await self.txn_repo.list_by_import_id_paginated(db, imp.id, page, page_size)
        total_pages = (total + page_size - 1) // page_size if total > 0 else 1

        txn_responses = [
            BankTransactionResponse(
                public_id=txn.public_id,
                transaction_at=txn.transaction_at,
                amount_inr=txn.amount_inr,
                direction=txn.direction,
                reference_id=txn.reference_id,
                utr=txn.utr,
                counterparty_name=txn.counterparty_name,
                description=txn.description,
                source=txn.source,
                source_transaction_key=txn.source_transaction_key,
                created_at=txn.created_at,
            )
            for txn in items
        ]

        return BankTransactionListResponse(
            items=txn_responses,
            total=total,
            page=page,
            page_size=page_size,
            total_pages=total_pages,
        )
