"""Statement parser service for inspecting and normalizing CSV and XLSX bank statements."""

import csv
import io
import re
import hashlib
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from typing import Any, Dict, List, Optional, Tuple
import openpyxl

from app.schemas.statement_import import HeaderItem, StatementColumnMapping


class StatementParserService:
    """Service to parse, validate, and normalize bank statement files."""

    MAX_FILE_SIZE: int = 5 * 1024 * 1024  # 5 MB
    ALLOWED_EXTENSIONS: set = {".csv", ".xlsx"}

    def validate_file_meta(self, filename: str, file_size: int) -> str:
        """Validate file size and extension. Rejects macros (.xlsm) and unsupported files."""
        if file_size > self.MAX_FILE_SIZE:
            raise ValueError(f"File size exceeds maximum allowed limit of {self.MAX_FILE_SIZE // (1024 * 1024)} MB.")

        lower_name = filename.lower()
        if lower_name.endswith(".csv"):
            return "csv"
        elif lower_name.endswith(".xlsx"):
            return "xlsx"
        else:
            raise ValueError("Unsupported file format. Only .csv and .xlsx files are supported.")

    def calculate_checksum(self, file_bytes: bytes) -> str:
        """Calculate SHA-256 checksum of raw file bytes."""
        return hashlib.sha256(file_bytes).hexdigest()

    def inspect_file(
        self,
        file_bytes: bytes,
        filename: str,
        sheet_name: Optional[str] = None,
        header_row_index: int = 1,
    ) -> Tuple[str, List[str], Optional[str], List[HeaderItem], List[Dict[str, Any]], int]:
        """Inspect file content, detect worksheets, extract headers at header_row_index, and return preview rows.

        Returns:
            (file_type, available_sheets, selected_sheet_name, headers, preview_rows, total_detected_rows)
        """
        file_type = self.validate_file_meta(filename, len(file_bytes))

        if file_type == "csv":
            available_sheets: List[str] = []
            selected_sheet_name = None
            headers, preview_rows, total_rows = self._inspect_csv(file_bytes, header_row_index)
        else:
            available_sheets, selected_sheet_name, headers, preview_rows, total_rows = self._inspect_xlsx(
                file_bytes, sheet_name, header_row_index
            )

        return file_type, available_sheets, selected_sheet_name, headers, preview_rows, total_rows

    def _inspect_csv(
        self, file_bytes: bytes, header_row_index: int
    ) -> Tuple[List[HeaderItem], List[Dict[str, Any]], int]:
        """Inspect CSV content with UTF-8 / UTF-8 BOM handling."""
        try:
            content_str = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                content_str = file_bytes.decode("latin-1")
            except Exception as exc:
                raise ValueError("Unable to decode CSV file. Ensure it is UTF-8 encoded.") from exc

        reader = csv.reader(io.StringIO(content_str))
        all_rows = list(reader)

        if not all_rows or len(all_rows) < header_row_index:
            raise ValueError(f"CSV file has insufficient rows (less than header row index {header_row_index}).")

        header_row = all_rows[header_row_index - 1]
        headers = [
            HeaderItem(column_index=idx, header=col_name.strip() if col_name else f"Column_{idx + 1}")
            for idx, col_name in enumerate(header_row)
        ]

        data_rows = all_rows[header_row_index:]
        total_rows = len(data_rows)
        preview_limit = min(20, total_rows)

        preview_rows: List[Dict[str, Any]] = []
        for row in data_rows[:preview_limit]:
            row_dict = {}
            for idx, item in enumerate(headers):
                val = row[idx] if idx < len(row) else ""
                row_dict[item.header] = val
            preview_rows.append(row_dict)

        return headers, preview_rows, total_rows

    def _inspect_xlsx(
        self, file_bytes: bytes, sheet_name: Optional[str], header_row_index: int
    ) -> Tuple[List[str], str, List[HeaderItem], List[Dict[str, Any]], int]:
        """Inspect XLSX workbook using openpyxl in data_only read_only mode."""
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError("Unable to parse Excel file. Ensure it is a valid .xlsx workbook.") from exc

        available_sheets = wb.sheetnames
        if not available_sheets:
            raise ValueError("Excel workbook contains no readable worksheets.")

        target_sheet = sheet_name if sheet_name and sheet_name in available_sheets else available_sheets[0]
        ws = wb[target_sheet]

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows or len(rows) < header_row_index:
            raise ValueError(f"Worksheet '{target_sheet}' has insufficient rows (less than header row index {header_row_index}).")

        header_row = rows[header_row_index - 1]
        headers = [
            HeaderItem(column_index=idx, header=str(col_name).strip() if col_name is not None else f"Column_{idx + 1}")
            for idx, col_name in enumerate(header_row)
        ]

        data_rows = rows[header_row_index:]
        total_rows = len(data_rows)
        preview_limit = min(20, total_rows)

        preview_rows: List[Dict[str, Any]] = []
        for row in data_rows[:preview_limit]:
            row_dict = {}
            for idx, item in enumerate(headers):
                val = row[idx] if idx < len(row) else ""
                row_dict[item.header] = str(val) if val is not None else ""
            preview_rows.append(row_dict)

        return available_sheets, target_sheet, headers, preview_rows, total_rows

    # Normalization Helpers
    def parse_amount(self, raw_val: Any) -> Optional[int]:
        """Parse whole-rupee amount from string or numeric input. Rejects fractional paisa values."""
        if raw_val is None:
            return None

        val_str = str(raw_val).strip()
        if not val_str:
            return None

        # Remove currency signs, commas, and whitespace
        cleaned = re.sub(r"[₹RsRSinrINR,\s]", "", val_str)

        try:
            num = float(cleaned)
        except ValueError as exc:
            raise ValueError(f"Invalid monetary amount: '{raw_val}'") from exc

        abs_num = abs(num)
        # Check whole rupee
        if round(abs_num) != round(abs_num, 2) or abs_num % 1 != 0:
            # Check if it has decimals like .00 vs .50
            if abs_num.is_integer():
                return int(abs_num)
            else:
                raise ValueError(f"Amount '{raw_val}' contains fractional paisa. System operates in whole Indian Rupees (INR).")

        return int(abs_num)

    def parse_date(self, raw_val: Any, default_tz_str: str = "Asia/Kolkata") -> Optional[datetime]:
        """Parse timestamp and normalize to timezone-aware UTC datetime."""
        if raw_val is None or str(raw_val).strip() == "":
            return None

        dt: Optional[datetime] = None

        if isinstance(raw_val, datetime):
            dt = raw_val
        else:
            val_str = str(raw_val).strip()
            # Try standard datetime formats
            date_formats = [
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%d/%m/%Y %H:%M:%S",
                "%d/%m/%Y %H:%M",
                "%d-%m-%Y %H:%M:%S",
                "%d-%m-%Y %H:%M",
                "%Y-%m-%d",
                "%d/%m/%Y",
                "%d-%m-%Y",
                "%Y/%m/%d",
            ]
            for fmt in date_formats:
                try:
                    dt = datetime.strptime(val_str, fmt)
                    break
                except ValueError:
                    continue

            if dt is None:
                # Attempt ISO format parse
                try:
                    dt = datetime.fromisoformat(val_str)
                except ValueError:
                    return None

        if dt is None:
            return None

        if dt.tzinfo is None:
            tz = ZoneInfo(default_tz_str)
            dt = dt.replace(tzinfo=tz)

        return dt.astimezone(timezone.utc)

    def parse_direction(self, raw_val: Any) -> Optional[str]:
        """Normalize transaction direction (CREDIT / DEBIT)."""
        if raw_val is None:
            return None

        val_str = str(raw_val).strip().upper()
        if not val_str:
            return None

        if any(keyword in val_str for keyword in ["CR", "CREDIT", "RECEIVED", "IN", "DEPOSIT"]):
            return "CREDIT"
        elif any(keyword in val_str for keyword in ["DR", "DEBIT", "SENT", "OUT", "WITHDRAWAL"]):
            return "DEBIT"

        return None

    def normalize_reference_id(self, raw_val: Any) -> Optional[str]:
        """Normalize payment reference ID: trim leading/trailing whitespace, preserve case and underscores."""
        if raw_val is None:
            return None

        val_str = str(raw_val).strip()
        return val_str if val_str else None

    def compute_source_transaction_key(
        self,
        source: str,
        reference_id: Optional[str],
        utr: Optional[str],
        amount_inr: Optional[int],
        direction: Optional[str],
        transaction_at: Optional[datetime],
        counterparty: Optional[str],
        description: Optional[str],
    ) -> str:
        """Formulate deterministic source_transaction_key according to priority hierarchy.

        File-Independent Fingerprint Rule:
        Excludes file_checksum, sheet_name, and row_index to ensure overlapping statement files
        deduplicate identical transactions cleanly regardless of file context.
        """
        # Priority 1: UTR if present
        if utr and utr.strip():
            return f"{source}_UTR_{utr.strip()}"

        # Priority 2: System Reference Code if present
        if reference_id and reference_id.strip():
            return f"{source}_REF_{reference_id.strip()}"

        # Priority 3: File-independent normalized content fingerprint
        dt_str = transaction_at.isoformat() if transaction_at else ""
        cp_str = (counterparty or "").strip().lower()
        desc_str = (description or "").strip().lower()
        raw_key = f"{source}|{reference_id or ''}|{amount_inr or ''}|{direction or ''}|{dt_str}|{cp_str}|{desc_str}"

        hash_digest = hashlib.sha256(raw_key.encode("utf-8")).hexdigest()
        return f"{source}_FPT_{hash_digest[:32]}"
