"""Script to generate synthetic multi-sheet Excel workbook fixture for testing."""

import os
import openpyxl

def generate_xlsx_fixture():
    wb = openpyxl.Workbook()
    
    # Sheet 1: Summary (Non-transaction metadata)
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Bank Account Statement Summary"])
    ws1.append(["Generated Date", "2026-08-17"])
    ws1.append(["Total Transactions", 4])
    ws1.append([])

    # Sheet 2: Transactions (Actual transaction rows)
    ws2 = wb.create_sheet(title="Transactions")
    ws2.append(["REPORT METADATA: INTERNAL CONFIDENTIAL"])  # Row 1: Header metadata
    ws2.append(["Date", "Type", "Transaction Remarks", "Amount", "UTR", "Payer Name", "Description"])  # Row 2: Headers
    ws2.append(["2026-08-16 10:30:00", "CREDIT", "ADITYA_3210_YON2", 2500, "987654321098", "Aditya Nair", "UPI enrollment fee"])
    ws2.append(["2026-08-16 11:15:00", "CREDIT", "VIKRAM_6655_A7K2", 4000, "987654321099", "Vikram Singh", "UPI course fee"])
    ws2.append(["2026-08-16 12:00:00", "CREDIT", "", 1500, "987654321100", "Rahul Verma", "Direct transfer without ref"])
    ws2.append(["2026-08-16 14:20:00", "DEBIT", "REF_OUT_100", 500, "987654321101", "Vendor", "Refund payment"])

    # Sheet 3: Account Details (Non-transaction metadata)
    ws3 = wb.create_sheet(title="Account Details")
    ws3.append(["Account Name", "Samagra Learning Solutions"])
    ws3.append(["Account Number", "91827364501928"])
    ws3.append(["IFSC Code", "SBIN0001234"])

    output_path = os.path.join(os.path.dirname(__file__), "sample_statement.xlsx")
    wb.save(output_path)
    print(f"Successfully generated {output_path}")

if __name__ == "__main__":
    generate_xlsx_fixture()
